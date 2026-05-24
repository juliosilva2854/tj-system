from fastapi import APIRouter, Depends
from fastapi.responses import Response
from database import db
from deps import require_roles
from permissions import CAN_VIEW_AUDIT, get_user_warehouse_scope, ADMIN_ROLES

router = APIRouter(tags=["audit"])

async def _scoped_query(user: dict) -> dict:
    q = {}
    if user['role'] != 'master':
        q['tenant_id'] = user.get('tenant_id', '')
    # gerentes veem apenas registros do seu escopo (warehouse_id)
    if user['role'] not in ADMIN_ROLES:
        scope = await get_user_warehouse_scope(user)
        if scope is not None:
            ids = list(scope) or ["__none__"]
            q['$or'] = [
                {"warehouse_id": {"$in": ids}},
                {"warehouse_id": "", "entity_type": {"$in": ["produto", "fornecedor", "usuario", "nota_fiscal"]}},
            ]
    return q

@router.get("/audit")
async def list_audit(user: dict = Depends(require_roles(*CAN_VIEW_AUDIT))):
    q = await _scoped_query(user)
    return await db.audit_logs.find(q, {"_id": 0}).sort('timestamp', -1).to_list(1000)

@router.get("/audit/export")
async def export_audit(user: dict = Depends(require_roles(*CAN_VIEW_AUDIT))):
    q = await _scoped_query(user)
    logs = await db.audit_logs.find(q, {"_id": 0}).sort('timestamp', -1).to_list(10000)
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    import io
    wb = Workbook(); ws = wb.active; ws.title = "Auditoria"
    for i, h in enumerate(['Data', 'Usuario', 'Acao', 'Entidade', 'ID', 'Detalhes'], 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
    for r, l in enumerate(logs, 2):
        ws.cell(row=r, column=1, value=l.get('timestamp', ''))
        ws.cell(row=r, column=2, value=l.get('user_email', ''))
        ws.cell(row=r, column=3, value=l.get('action', ''))
        ws.cell(row=r, column=4, value=l.get('entity_type', ''))
        ws.cell(row=r, column=5, value=l.get('entity_id', ''))
        ws.cell(row=r, column=6, value=str(l.get('changes', '') or ''))
    for col in 'ABCDEF':
        ws.column_dimensions[col].width = 22
    buf = io.BytesIO(); wb.save(buf)
    return Response(content=buf.getvalue(),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": "attachment; filename=auditoria.xlsx"})
