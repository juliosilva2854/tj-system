from fastapi import FastAPI, APIRouter, Depends, HTTPException, status, Header, Request, UploadFile, File
from fastapi.responses import Response
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from slowapi import Limiter, _rate_limit_exceeded_handler
from slowapi.util import get_remote_address
from slowapi.errors import RateLimitExceeded
import os, logging, json, uuid, base64
from pathlib import Path
from datetime import datetime, timezone
from typing import Optional, List

from models import (
    TenantCreate, Tenant, UserCreate, UserLogin, UserOut,
    WarehouseCreate, Warehouse, ProductCreate, Product,
    InvoiceCreate, Invoice, InvoiceItemInput, OCRRequest,
    RequisitionCreate, Requisition, SaleCreate,
    SupplierCreate, gen_id
)
from auth import hash_password, verify_password, create_access_token, create_refresh_token, decode_token
from audit import AuditLogger
from nfe_parser import parse_nfe_xml
from report_export import generate_financial_pdf, generate_financial_excel
from email_service import send_email, build_stock_alert_email

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]
audit = AuditLogger(db)

limiter = Limiter(key_func=get_remote_address)
app = FastAPI(title="Gestao TJ - SaaS Multi-Tenant")
app.state.limiter = limiter
app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)

api = APIRouter(prefix="/api")
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ═══════════════════════════════════════════════
# SECURITY MIDDLEWARE
# ═══════════════════════════════════════════════

async def get_current_user(authorization: Optional[str] = Header(None)) -> dict:
    if not authorization or not authorization.startswith('Bearer '):
        raise HTTPException(status_code=401, detail="Nao autenticado")
    payload = decode_token(authorization[7:])
    if not payload or payload.get('type') != 'access':
        raise HTTPException(status_code=401, detail="Token invalido ou expirado")
    return payload

def require_roles(*roles):
    async def checker(user: dict = Depends(get_current_user)):
        if user['role'] not in roles:
            raise HTTPException(status_code=403, detail="Permissao insuficiente")
        return user
    return checker

async def verify_tenant_access(user: dict, tenant_id: str):
    if user['role'] == 'master':
        return
    if user.get('tenant_id') != tenant_id:
        raise HTTPException(status_code=403, detail="Acesso negado a este estabelecimento")

async def get_user_tenant(user: dict) -> str:
    if user['role'] == 'master':
        return None
    tid = user.get('tenant_id')
    if not tid:
        raise HTTPException(status_code=403, detail="Usuario sem estabelecimento vinculado")
    return tid

# ═══════════════════════════════════════════════
# AUTH
# ═══════════════════════════════════════════════

@api.post("/auth/login")
@limiter.limit("10/minute")
async def login(request: Request, creds: UserLogin):
    doc = await db.users.find_one({"email": creds.email}, {"_id": 0})
    if not doc or not verify_password(creds.password, doc['password_hash']):
        raise HTTPException(status_code=401, detail="Email ou senha incorretos")
    if not doc.get('active', True):
        raise HTTPException(status_code=403, detail="Conta inativa")
    access = create_access_token(doc['id'], doc['email'], doc['role'], doc.get('tenant_id', ''))
    refresh = create_refresh_token(doc['id'])
    user_out = {k: doc[k] for k in ['id', 'email', 'name', 'role', 'tenant_id', 'warehouse_id', 'active', 'created_at'] if k in doc}
    return {"access_token": access, "refresh_token": refresh, "user": user_out}

@api.post("/auth/refresh")
async def refresh_token(request: Request):
    body = await request.json()
    token = body.get('refresh_token')
    if not token:
        raise HTTPException(status_code=400, detail="Refresh token obrigatorio")
    payload = decode_token(token)
    if not payload or payload.get('type') != 'refresh':
        raise HTTPException(status_code=401, detail="Refresh token invalido")
    doc = await db.users.find_one({"id": payload['sub']}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=401, detail="Usuario nao encontrado")
    access = create_access_token(doc['id'], doc['email'], doc['role'], doc.get('tenant_id', ''))
    return {"access_token": access}

@api.post("/auth/register")
async def register(data: UserCreate, user: dict = Depends(require_roles("master", "admin"))):
    if user['role'] == 'admin' and data.role == 'master':
        raise HTTPException(status_code=403, detail="Admin nao pode criar master")
    if user['role'] == 'admin':
        data.tenant_id = user.get('tenant_id')
    existing = await db.users.find_one({"email": data.email}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="Email ja cadastrado")
    now = datetime.now(timezone.utc).isoformat()
    doc = {
        "id": gen_id(), "email": data.email, "name": data.name, "role": data.role,
        "tenant_id": data.tenant_id or "", "warehouse_id": data.warehouse_id or "",
        "password_hash": hash_password(data.password), "active": True, "created_at": now
    }
    await db.users.insert_one(doc)
    await audit.log(user['sub'], user['email'], "CRIAR", "usuario", doc['id'], doc.get('tenant_id', ''))
    doc.pop('password_hash')
    return doc

# ═══════════════════════════════════════════════
# TENANTS
# ═══════════════════════════════════════════════

@api.post("/tenants")
async def create_tenant(data: TenantCreate, user: dict = Depends(require_roles("master"))):
    existing = await db.tenants.find_one({"slug": data.slug}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="Slug ja em uso")
    now = datetime.now(timezone.utc).isoformat()
    doc = {"id": gen_id(), "name": data.name, "slug": data.slug, "active": True, "created_at": now}
    await db.tenants.insert_one(doc)
    await audit.log(user['sub'], user['email'], "CRIAR", "tenant", doc['id'])
    return doc

@api.get("/tenants")
async def list_tenants(user: dict = Depends(require_roles("master"))):
    docs = await db.tenants.find({}, {"_id": 0}).to_list(1000)
    return docs

# ═══════════════════════════════════════════════
# USERS
# ═══════════════════════════════════════════════

@api.get("/users")
async def list_users(user: dict = Depends(require_roles("master", "admin"))):
    q = {}
    if user['role'] == 'admin':
        q['tenant_id'] = user.get('tenant_id')
    docs = await db.users.find(q, {"_id": 0, "password_hash": 0}).to_list(1000)
    return docs

@api.patch("/users/{uid}")
async def update_user(uid: str, request: Request, user: dict = Depends(require_roles("master", "admin"))):
    body = await request.json()
    body.pop('id', None)
    body.pop('email', None)
    if 'password' in body:
        body['password_hash'] = hash_password(body.pop('password'))
    target = await db.users.find_one({"id": uid}, {"_id": 0})
    if not target:
        raise HTTPException(status_code=404, detail="Usuario nao encontrado")
    if user['role'] == 'admin' and target.get('tenant_id') != user.get('tenant_id'):
        raise HTTPException(status_code=403, detail="Sem permissao")
    await db.users.update_one({"id": uid}, {"$set": body})
    await audit.log(user['sub'], user['email'], "EDITAR", "usuario", uid, user.get('tenant_id', ''))
    return {"message": "Atualizado"}

@api.delete("/users/{uid}")
async def delete_user(uid: str, user: dict = Depends(require_roles("master"))):
    if uid == user['sub']:
        raise HTTPException(status_code=400, detail="Nao pode excluir a si mesmo")
    r = await db.users.delete_one({"id": uid})
    if r.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Nao encontrado")
    return {"message": "Excluido"}

# ═══════════════════════════════════════════════
# WAREHOUSES (PAI / FILHO)
# ═══════════════════════════════════════════════

@api.post("/warehouses")
async def create_warehouse(data: WarehouseCreate, user: dict = Depends(require_roles("master", "admin"))):
    tid = user.get('tenant_id', '')
    if user['role'] == 'master' and not tid:
        body = data.model_dump()
        tid = body.get('tenant_id', '')
    now = datetime.now(timezone.utc).isoformat()
    doc = {
        "id": gen_id(), "tenant_id": tid, "name": data.name, "location": data.location,
        "description": data.description or "", "type": data.type,
        "parent_id": data.parent_id or "", "sectors": data.sectors,
        "active": True, "created_at": now, "created_by": user['sub']
    }
    await db.warehouses.insert_one(doc)
    await audit.log(user['sub'], user['email'], "CRIAR", "deposito", doc['id'], tid)
    return doc

@api.get("/warehouses")
async def list_warehouses(user: dict = Depends(get_current_user)):
    q = {}
    tid = user.get('tenant_id')
    if user['role'] != 'master' and tid:
        q['tenant_id'] = tid
    if user['role'] == 'operacional' and user.get('warehouse_id'):
        q['id'] = user.get('warehouse_id')
    docs = await db.warehouses.find(q, {"_id": 0}).to_list(1000)
    return docs

@api.patch("/warehouses/{wid}")
async def update_warehouse(wid: str, request: Request, user: dict = Depends(require_roles("master", "admin"))):
    body = await request.json()
    target = await db.warehouses.find_one({"id": wid}, {"_id": 0})
    if not target:
        raise HTTPException(status_code=404, detail="Nao encontrado")
    if user['role'] != 'master':
        await verify_tenant_access(user, target['tenant_id'])
    body.pop('id', None)
    body.pop('tenant_id', None)
    await db.warehouses.update_one({"id": wid}, {"$set": body})
    return {"message": "Atualizado"}

@api.delete("/warehouses/{wid}")
async def delete_warehouse(wid: str, user: dict = Depends(require_roles("master", "admin"))):
    target = await db.warehouses.find_one({"id": wid}, {"_id": 0})
    if not target:
        raise HTTPException(status_code=404, detail="Nao encontrado")
    if user['role'] != 'master':
        await verify_tenant_access(user, target['tenant_id'])
    await db.warehouses.delete_one({"id": wid})
    return {"message": "Excluido"}

# ═══════════════════════════════════════════════
# PRODUCTS
# ═══════════════════════════════════════════════

@api.post("/products")
async def create_product(data: ProductCreate, user: dict = Depends(get_current_user)):
    tid = user.get('tenant_id', '')
    now = datetime.now(timezone.utc).isoformat()
    doc = {
        "id": gen_id(), "tenant_id": tid, **data.model_dump(),
        "available_qty": 0, "active": True, "created_at": now, "created_by": user['sub']
    }
    await db.products.insert_one(doc)
    return doc

@api.get("/products")
async def list_products(user: dict = Depends(get_current_user)):
    q = {}
    if user['role'] != 'master':
        q['tenant_id'] = user.get('tenant_id', '')
    docs = await db.products.find(q, {"_id": 0}).to_list(5000)
    return docs

@api.patch("/products/{pid}")
async def update_product(pid: str, request: Request, user: dict = Depends(get_current_user)):
    body = await request.json()
    target = await db.products.find_one({"id": pid}, {"_id": 0})
    if not target:
        raise HTTPException(status_code=404, detail="Nao encontrado")
    if user['role'] != 'master':
        await verify_tenant_access(user, target['tenant_id'])
    body.pop('id', None)
    body.pop('tenant_id', None)
    await db.products.update_one({"id": pid}, {"$set": body})
    return {"message": "Atualizado"}

@api.post("/products/{pid}/transfer")
async def transfer_product(pid: str, warehouse_id: str, quantity: float, sector: str = "", user: dict = Depends(get_current_user)):
    product = await db.products.find_one({"id": pid}, {"_id": 0})
    if not product:
        raise HTTPException(status_code=404, detail="Produto nao encontrado")
    if user['role'] != 'master':
        await verify_tenant_access(user, product['tenant_id'])
    wh = await db.warehouses.find_one({"id": warehouse_id}, {"_id": 0})
    if not wh:
        raise HTTPException(status_code=404, detail="Deposito nao encontrado")
    avail = product.get('available_qty', 0)
    if quantity > avail:
        quantity = avail
    if quantity <= 0:
        raise HTTPException(status_code=400, detail="Quantidade invalida")
    inv = await db.inventory.find_one({"product_id": pid, "warehouse_id": warehouse_id, "tenant_id": product['tenant_id']}, {"_id": 0})
    now = datetime.now(timezone.utc).isoformat()
    if inv:
        await db.inventory.update_one({"id": inv['id']}, {"$set": {"quantity": inv['quantity'] + quantity, "updated_at": now}})
    else:
        await db.inventory.insert_one({
            "id": gen_id(), "tenant_id": product['tenant_id'], "product_id": pid, "warehouse_id": warehouse_id,
            "quantity": quantity, "updated_at": now
        })
    new_avail = avail - quantity
    if new_avail <= 0:
        await db.products.delete_one({"id": pid})
        msg = f"Transferido {quantity} para {wh['name']}. Produto removido da aba."
    else:
        await db.products.update_one({"id": pid}, {"$set": {"available_qty": new_avail}})
        msg = f"Transferido {quantity} para {wh['name']}. Restam {new_avail}."
    await audit.log(user['sub'], user['email'], "TRANSFERIR", "produto", pid, product['tenant_id'], {"deposito": wh['name'], "quantidade": quantity, "setor": sector})
    return {"message": msg, "removed": new_avail <= 0}

# ═══════════════════════════════════════════════
# INVENTORY
# ═══════════════════════════════════════════════

@api.get("/inventory")
async def list_inventory(user: dict = Depends(get_current_user)):
    q = {}
    if user['role'] != 'master':
        q['tenant_id'] = user.get('tenant_id', '')
    if user['role'] == 'operacional' and user.get('warehouse_id'):
        q['warehouse_id'] = user.get('warehouse_id')
    items = await db.inventory.find(q, {"_id": 0}).to_list(5000)
    result = []
    for it in items:
        p = await db.products.find_one({"id": it['product_id']}, {"_id": 0}) or await db.inventory.find_one({"product_id": it['product_id']}, {"_id": 0})
        w = await db.warehouses.find_one({"id": it['warehouse_id']}, {"_id": 0})
        pdata = await db.products.find_one({"id": it['product_id']}, {"_id": 0})
        pname = pdata['name'] if pdata else it.get('product_name', 'Desconhecido')
        psku = pdata['sku'] if pdata else it.get('product_sku', '')
        min_s = pdata.get('min_stock', 0) if pdata else 0
        result.append({
            "id": it['id'], "tenant_id": it.get('tenant_id', ''),
            "product_id": it['product_id'], "product_name": pname, "product_sku": psku,
            "warehouse_id": it['warehouse_id'], "warehouse_name": w['name'] if w else 'Desconhecido',
            "warehouse_type": w.get('type', 'pai') if w else 'pai',
            "quantity": it['quantity'], "min_stock": min_s,
            "updated_at": it.get('updated_at', '')
        })
    return result

@api.post("/inventory/adjust")
async def adjust_inventory(product_id: str, warehouse_id: str, quantity: float, user: dict = Depends(get_current_user)):
    tid = user.get('tenant_id', '')
    inv = await db.inventory.find_one({"product_id": product_id, "warehouse_id": warehouse_id}, {"_id": 0})
    now = datetime.now(timezone.utc).isoformat()
    if inv:
        new_qty = max(0, inv['quantity'] + quantity)
        await db.inventory.update_one({"id": inv['id']}, {"$set": {"quantity": new_qty, "updated_at": now}})
    else:
        if quantity < 0:
            raise HTTPException(status_code=400, detail="Nao ha estoque para dar baixa")
        await db.inventory.insert_one({
            "id": gen_id(), "tenant_id": tid, "product_id": product_id,
            "warehouse_id": warehouse_id, "quantity": max(0, quantity), "updated_at": now
        })
    await audit.log(user['sub'], user['email'], "AJUSTAR", "estoque", f"{product_id}:{warehouse_id}", tid, {"quantidade": quantity})
    return {"message": "Estoque ajustado"}

# ═══════════════════════════════════════════════
# REQUISITIONS (FILHO -> PAI)
# ═══════════════════════════════════════════════

@api.post("/requisitions")
async def create_requisition(data: RequisitionCreate, user: dict = Depends(get_current_user)):
    tid = user.get('tenant_id', '')
    wid = user.get('warehouse_id', '')
    if not wid:
        raise HTTPException(status_code=400, detail="Voce precisa estar vinculado a um deposito")
    wh = await db.warehouses.find_one({"id": wid}, {"_id": 0})
    if not wh or wh.get('type') != 'filho':
        raise HTTPException(status_code=400, detail="Requisicoes sao criadas apenas por depositos filhos")
    parent_id = wh.get('parent_id', '')
    if not parent_id:
        raise HTTPException(status_code=400, detail="Deposito filho sem pai vinculado")
    now = datetime.now(timezone.utc).isoformat()
    doc = {
        "id": gen_id(), "tenant_id": tid, "from_warehouse_id": wid, "to_warehouse_id": parent_id,
        "items": [i.model_dump() for i in data.items], "notes": data.notes or "",
        "status": "pending", "created_at": now, "created_by": user['sub']
    }
    await db.requisitions.insert_one(doc)
    await audit.log(user['sub'], user['email'], "CRIAR", "requisicao", doc['id'], tid)
    return doc

@api.get("/requisitions")
async def list_requisitions(user: dict = Depends(get_current_user)):
    q = {}
    tid = user.get('tenant_id', '')
    if user['role'] != 'master':
        q['tenant_id'] = tid
    if user['role'] == 'operacional':
        q['from_warehouse_id'] = user.get('warehouse_id', '')
    docs = await db.requisitions.find(q, {"_id": 0}).sort('created_at', -1).to_list(1000)
    return docs

@api.post("/requisitions/{rid}/approve")
async def approve_requisition(rid: str, user: dict = Depends(require_roles("master", "admin", "logistica"))):
    req = await db.requisitions.find_one({"id": rid}, {"_id": 0})
    if not req:
        raise HTTPException(status_code=404, detail="Requisicao nao encontrada")
    if user['role'] != 'master':
        await verify_tenant_access(user, req['tenant_id'])
    if req['status'] != 'pending':
        raise HTTPException(status_code=400, detail="Requisicao ja processada")
    now = datetime.now(timezone.utc).isoformat()
    pai_id = req['to_warehouse_id']
    filho_id = req['from_warehouse_id']
    for item in req['items']:
        pid = item['product_id']
        qty = item['quantity']
        pai_inv = await db.inventory.find_one({"product_id": pid, "warehouse_id": pai_id}, {"_id": 0})
        if not pai_inv or pai_inv['quantity'] < qty:
            available = pai_inv['quantity'] if pai_inv else 0
            raise HTTPException(status_code=400, detail=f"Estoque insuficiente no almoxarifado para {item['product_name']}. Disponivel: {available}")
        new_pai_qty = max(0, pai_inv['quantity'] - qty)
        await db.inventory.update_one({"id": pai_inv['id']}, {"$set": {"quantity": new_pai_qty, "updated_at": now}})
        filho_inv = await db.inventory.find_one({"product_id": pid, "warehouse_id": filho_id}, {"_id": 0})
        if filho_inv:
            await db.inventory.update_one({"id": filho_inv['id']}, {"$set": {"quantity": filho_inv['quantity'] + qty, "updated_at": now}})
        else:
            await db.inventory.insert_one({
                "id": gen_id(), "tenant_id": req['tenant_id'], "product_id": pid,
                "warehouse_id": filho_id, "quantity": qty, "updated_at": now
            })
    await db.requisitions.update_one({"id": rid}, {"$set": {"status": "approved", "resolved_at": now, "resolved_by": user['sub']}})
    await audit.log(user['sub'], user['email'], "APROVAR", "requisicao", rid, req['tenant_id'])
    return {"message": "Requisicao aprovada. Itens transferidos."}

@api.post("/requisitions/{rid}/reject")
async def reject_requisition(rid: str, user: dict = Depends(require_roles("master", "admin", "logistica"))):
    req = await db.requisitions.find_one({"id": rid}, {"_id": 0})
    if not req:
        raise HTTPException(status_code=404, detail="Nao encontrada")
    if user['role'] != 'master':
        await verify_tenant_access(user, req['tenant_id'])
    now = datetime.now(timezone.utc).isoformat()
    await db.requisitions.update_one({"id": rid}, {"$set": {"status": "rejected", "resolved_at": now, "resolved_by": user['sub']}})
    return {"message": "Requisicao rejeitada"}

# ═══════════════════════════════════════════════
# SUPPLIERS
# ═══════════════════════════════════════════════

@api.post("/suppliers")
async def create_supplier(data: SupplierCreate, user: dict = Depends(get_current_user)):
    tid = user.get('tenant_id', '')
    now = datetime.now(timezone.utc).isoformat()
    doc = {"id": gen_id(), "tenant_id": tid, **data.model_dump(), "active": True, "created_at": now, "created_by": user['sub']}
    await db.suppliers.insert_one(doc)
    return doc

@api.get("/suppliers")
async def list_suppliers(user: dict = Depends(get_current_user)):
    q = {}
    if user['role'] != 'master':
        q['tenant_id'] = user.get('tenant_id', '')
    return await db.suppliers.find(q, {"_id": 0}).to_list(1000)

@api.patch("/suppliers/{sid}")
async def update_supplier(sid: str, request: Request, user: dict = Depends(get_current_user)):
    body = await request.json()
    target = await db.suppliers.find_one({"id": sid}, {"_id": 0})
    if not target:
        raise HTTPException(status_code=404, detail="Nao encontrado")
    if user['role'] != 'master':
        await verify_tenant_access(user, target['tenant_id'])
    body.pop('id', None)
    body.pop('tenant_id', None)
    await db.suppliers.update_one({"id": sid}, {"$set": body})
    return {"message": "Atualizado"}

@api.delete("/suppliers/{sid}")
async def delete_supplier(sid: str, user: dict = Depends(get_current_user)):
    target = await db.suppliers.find_one({"id": sid}, {"_id": 0})
    if not target:
        raise HTTPException(status_code=404, detail="Nao encontrado")
    if user['role'] != 'master':
        await verify_tenant_access(user, target['tenant_id'])
    await db.suppliers.delete_one({"id": sid})
    return {"message": "Excluido"}

# ═══════════════════════════════════════════════
# INVOICES + OCR (Gemini)
# ═══════════════════════════════════════════════

@api.post("/invoices")
async def create_invoice(data: InvoiceCreate, user: dict = Depends(get_current_user)):
    tid = user.get('tenant_id', '')
    now = datetime.now(timezone.utc).isoformat()
    doc = {
        "id": gen_id(), "tenant_id": tid, **data.model_dump(),
        "items": [i.model_dump() for i in data.items],
        "status": "pending", "type": "entrada", "created_at": now, "created_by": user['sub']
    }
    await db.invoices.insert_one(doc)
    return doc

@api.get("/invoices")
async def list_invoices(user: dict = Depends(get_current_user)):
    q = {}
    if user['role'] != 'master':
        q['tenant_id'] = user.get('tenant_id', '')
    return await db.invoices.find(q, {"_id": 0}).to_list(5000)

@api.post("/invoices/{iid}/process-items")
async def process_invoice_items(iid: str, user: dict = Depends(get_current_user)):
    inv = await db.invoices.find_one({"id": iid}, {"_id": 0})
    if not inv:
        raise HTTPException(status_code=404, detail="Nota nao encontrada")
    if user['role'] != 'master':
        await verify_tenant_access(user, inv['tenant_id'])
    tid = inv['tenant_id']
    created = 0
    for item in inv.get('items', []):
        pname = item.get('product_name', '')
        if not pname:
            continue
        sku = item.get('product_sku', '') or pname[:10].upper().replace(' ', '')
        qty = item.get('quantity', 0)
        existing = await db.products.find_one({"sku": sku, "tenant_id": tid}, {"_id": 0})
        now = datetime.now(timezone.utc).isoformat()
        if existing:
            new_avail = existing.get('available_qty', 0) + qty
            await db.products.update_one({"id": existing['id']}, {"$set": {"available_qty": new_avail}})
        else:
            await db.products.insert_one({
                "id": gen_id(), "tenant_id": tid, "name": pname, "sku": sku,
                "description": "", "category": "", "unit": "UN",
                "cost_price": item.get('unit_price', 0), "min_stock": 0,
                "available_qty": qty, "active": True, "created_at": now,
                "created_by": user['sub']
            })
            created += 1
    await db.invoices.update_one({"id": iid}, {"$set": {"status": "processed"}})
    await audit.log(user['sub'], user['email'], "PROCESSAR", "nota_fiscal", iid, tid, {"produtos_criados": created})
    return {"message": f"Itens enviados para Produtos. {created} novos criados.", "products_created": created}

@api.post("/invoices/ocr")
async def ocr_invoice(data: OCRRequest, user: dict = Depends(get_current_user)):
    import google.generativeai as genai
    api_key = os.environ.get('GEMINI_API_KEY')
    if not api_key:
        raise HTTPException(status_code=500, detail="GEMINI_API_KEY nao configurada")
    genai.configure(api_key=api_key)
    model = genai.GenerativeModel('gemini-2.0-flash')
    prompt = """Analise esta nota fiscal brasileira. Extraia com PRECISAO:
- QUANTIDADE = numero de unidades (NAO confunda com codigo)
- VALOR UNITARIO = preco de 1 unidade
Retorne SOMENTE JSON:
{"invoice_number":"","supplier_name":"","issue_date":"YYYY-MM-DD","total_value":0,"tax_value":0,"items":[{"product_name":"","product_sku":"","quantity":0,"unit_price":0,"total":0}]}"""
    import base64 as b64mod
    img_bytes = b64mod.b64decode(data.image_base64)
    response = model.generate_content([prompt, {"mime_type": "image/jpeg", "data": img_bytes}])
    txt = response.text.strip()
    if txt.startswith('```json'):
        txt = txt[7:]
    if txt.startswith('```'):
        txt = txt[3:]
    if txt.endswith('```'):
        txt = txt[:-3]
    return json.loads(txt.strip())

@api.post("/invoices/upload")
async def upload_invoice(file: UploadFile = File(...), user: dict = Depends(get_current_user)):
    content = await file.read()
    fname = (file.filename or '').lower()
    if fname.endswith('.xml'):
        parsed = parse_nfe_xml(content)
        return {"source": "xml", "data": parsed}
    elif fname.endswith('.pdf') or fname.endswith('.jpg') or fname.endswith('.jpeg') or fname.endswith('.png'):
        import google.generativeai as genai
        api_key = os.environ.get('GEMINI_API_KEY')
        if not api_key:
            raise HTTPException(status_code=500, detail="GEMINI_API_KEY nao configurada")
        genai.configure(api_key=api_key)
        model = genai.GenerativeModel('gemini-2.0-flash')
        prompt = 'Analise esta nota fiscal brasileira. QUANTIDADE = unidades compradas. Retorne SOMENTE JSON: {"invoice_number":"","supplier_name":"","issue_date":"YYYY-MM-DD","total_value":0,"tax_value":0,"items":[{"product_name":"","product_sku":"","quantity":0,"unit_price":0,"total":0}]}'
        mime = "application/pdf" if fname.endswith('.pdf') else "image/jpeg"
        response = model.generate_content([prompt, {"mime_type": mime, "data": content}])
        txt = response.text.strip()
        if txt.startswith('```json'):
            txt = txt[7:]
        if txt.startswith('```'):
            txt = txt[3:]
        if txt.endswith('```'):
            txt = txt[:-3]
        return {"source": "ocr", "data": json.loads(txt.strip())}
    raise HTTPException(status_code=400, detail="Use PDF, XML, JPG ou PNG")

# ═══════════════════════════════════════════════
# SALES
# ═══════════════════════════════════════════════

@api.post("/sales")
async def create_sale(data: SaleCreate, user: dict = Depends(get_current_user)):
    tid = user.get('tenant_id', '')
    now = datetime.now(timezone.utc).isoformat()
    last = await db.sales.find_one({"tenant_id": tid}, {"_id": 0}, sort=[('created_at', -1)])
    num = 1
    if last and 'sale_number' in last:
        try:
            num = int(last['sale_number'][3:]) + 1
        except ValueError:
            pass
    doc = {
        "id": gen_id(), "tenant_id": tid, "sale_number": f"VND{str(num).zfill(6)}",
        "warehouse_id": data.warehouse_id, "customer_name": data.customer_name or "",
        "items": [i.model_dump() for i in data.items],
        "subtotal": data.subtotal, "discount": data.discount, "total": data.total,
        "payment_method": data.payment_method or "", "status": "completed",
        "created_at": now, "created_by": user['sub']
    }
    await db.sales.insert_one(doc)
    for item in data.items:
        inv = await db.inventory.find_one({"product_id": item.product_id, "warehouse_id": data.warehouse_id}, {"_id": 0})
        if inv:
            new_qty = max(0, inv['quantity'] - item.quantity)
            await db.inventory.update_one({"id": inv['id']}, {"$set": {"quantity": new_qty, "updated_at": now}})
    await audit.log(user['sub'], user['email'], "CRIAR", "venda", doc['id'], tid)
    return doc

@api.get("/sales")
async def list_sales(user: dict = Depends(get_current_user)):
    q = {}
    if user['role'] != 'master':
        q['tenant_id'] = user.get('tenant_id', '')
    return await db.sales.find(q, {"_id": 0}).sort('created_at', -1).to_list(5000)

# ═══════════════════════════════════════════════
# DASHBOARD / REPORTS / AUDIT / NOTIFICATIONS
# ═══════════════════════════════════════════════

@api.get("/dashboard/stats")
async def dashboard_stats(user: dict = Depends(get_current_user)):
    q = {}
    if user['role'] != 'master':
        q['tenant_id'] = user.get('tenant_id', '')
    prods = await db.products.count_documents({**q, "active": True})
    suppliers = await db.suppliers.count_documents({**q, "active": True})
    whs = await db.warehouses.count_documents({**q, "active": True})
    pending = await db.invoices.count_documents({**q, "status": "pending"})
    pending_reqs = await db.requisitions.count_documents({**q, "status": "pending"})
    inv = await db.inventory.find(q, {"_id": 0}).to_list(5000)
    low = 0
    for it in inv:
        p = await db.products.find_one({"id": it['product_id']}, {"_id": 0})
        if p and p.get('min_stock', 0) > 0 and it['quantity'] <= p['min_stock']:
            low += 1
    return {
        "total_products": prods, "total_suppliers": suppliers, "total_warehouses": whs,
        "pending_invoices": pending, "pending_requisitions": pending_reqs,
        "low_stock_alerts": low
    }

@api.get("/dashboard/alerts")
async def dashboard_alerts(user: dict = Depends(get_current_user)):
    q = {}
    if user['role'] != 'master':
        q['tenant_id'] = user.get('tenant_id', '')
    inv = await db.inventory.find(q, {"_id": 0}).to_list(5000)
    alerts = []
    for it in inv:
        p = await db.products.find_one({"id": it['product_id']}, {"_id": 0})
        w = await db.warehouses.find_one({"id": it['warehouse_id']}, {"_id": 0})
        if p and w and p.get('min_stock', 0) > 0 and it['quantity'] <= p['min_stock']:
            alerts.append({"product_name": p['name'], "warehouse_name": w['name'], "current_quantity": it['quantity'], "min_stock": p['min_stock']})
    return alerts

@api.get("/reports/financial")
async def financial_report(period: str, user: dict = Depends(require_roles("master", "admin"))):
    q = {}
    if user['role'] != 'master':
        q['tenant_id'] = user.get('tenant_id', '')
    if period == "month":
        start = datetime.now(timezone.utc).replace(day=1, hour=0, minute=0, second=0).isoformat()
    else:
        start = datetime.now(timezone.utc).replace(month=1, day=1, hour=0, minute=0, second=0).isoformat()
    sales = await db.sales.find({**q, "created_at": {"$gte": start}, "status": "completed"}, {"_id": 0}).to_list(10000)
    revenue = sum(s.get('total', 0) for s in sales)
    cost = 0
    for s in sales:
        for it in s.get('items', []):
            p = await db.products.find_one({"id": it.get('product_id')}, {"_id": 0})
            if p:
                cost += p.get('cost_price', 0) * it.get('quantity', 0)
    gp = revenue - cost
    margin = (gp / revenue * 100) if revenue > 0 else 0
    return {"period": period, "revenue": revenue, "cost": cost, "gross_profit": gp, "profit_margin": margin, "expenses": 0, "net_profit": gp}

@api.get("/reports/abc-curve")
async def abc_curve(user: dict = Depends(require_roles("master", "admin"))):
    q = {"status": "completed"}
    if user['role'] != 'master':
        q['tenant_id'] = user.get('tenant_id', '')
    sales = await db.sales.find(q, {"_id": 0}).to_list(10000)
    pm = {}
    for s in sales:
        for it in s.get('items', []):
            pid = it.get('product_id', '')
            pm.setdefault(pid, {"product_name": it.get('product_name', ''), "revenue": 0, "quantity": 0})
            pm[pid]['revenue'] += it.get('total', 0)
            pm[pid]['quantity'] += it.get('quantity', 0)
    items = sorted(pm.values(), key=lambda x: x['revenue'], reverse=True)
    total = sum(i['revenue'] for i in items)
    cum = 0
    for i in items:
        cum += i['revenue']
        pct = (cum / total * 100) if total > 0 else 0
        i['percentage'] = round(i['revenue'] / total * 100, 1) if total > 0 else 0
        i['cumulative'] = round(pct, 1)
        i['class'] = 'A' if pct <= 80 else ('B' if pct <= 95 else 'C')
    return {"items": items, "total_revenue": total}

@api.get("/reports/inventory-turnover")
async def inventory_turnover(user: dict = Depends(require_roles("master", "admin"))):
    pq = {"active": True}
    sq = {"status": "completed"}
    if user['role'] != 'master':
        pq['tenant_id'] = user.get('tenant_id', '')
        sq['tenant_id'] = user.get('tenant_id', '')
    products = await db.products.find(pq, {"_id": 0}).to_list(5000)
    sales = await db.sales.find(sq, {"_id": 0}).to_list(10000)
    ps = {}
    for s in sales:
        for it in s.get('items', []):
            ps[it.get('product_id', '')] = ps.get(it.get('product_id', ''), 0) + it.get('quantity', 0)
    inv_q = {}
    if user['role'] != 'master':
        inv_q['tenant_id'] = user.get('tenant_id', '')
    inv_items = await db.inventory.find(inv_q, {"_id": 0}).to_list(5000)
    pstock = {}
    for i in inv_items:
        pstock[i['product_id']] = pstock.get(i['product_id'], 0) + i['quantity']
    results = []
    for p in products:
        sold = ps.get(p['id'], 0)
        stock = pstock.get(p['id'], 0)
        avg = stock if stock > 0 else 1
        turn = round(sold / avg, 2)
        days = round(avg / (sold / 30), 0) if sold > 0 else 999
        results.append({"product_name": p['name'], "sku": p['sku'], "total_sold": sold, "current_stock": stock, "turnover_rate": turn, "days_of_coverage": min(days, 999), "status": 'critico' if days < 7 else 'baixo' if days < 15 else 'normal' if days < 60 else 'excesso'})
    return {"items": sorted(results, key=lambda x: x['turnover_rate'], reverse=True)}

@api.get("/reports/export/pdf")
async def export_pdf(period: str, user: dict = Depends(require_roles("master", "admin"))):
    report = await financial_report(period, user)
    pdf = generate_financial_pdf(report, "Mes Atual" if period == "month" else "Ano Atual")
    return Response(content=pdf, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename=relatorio_{period}.pdf"})

@api.get("/reports/export/excel")
async def export_excel(period: str, user: dict = Depends(require_roles("master", "admin"))):
    report = await financial_report(period, user)
    excel = generate_financial_excel(report, "Mes Atual" if period == "month" else "Ano Atual")
    return Response(content=excel, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename=relatorio_{period}.xlsx"})

@api.get("/audit")
async def list_audit(user: dict = Depends(require_roles("master", "admin"))):
    q = {}
    if user['role'] != 'master':
        q['tenant_id'] = user.get('tenant_id', '')
    return await db.audit_logs.find(q, {"_id": 0}).sort('timestamp', -1).to_list(1000)

@api.get("/audit/export")
async def export_audit(user: dict = Depends(require_roles("master", "admin"))):
    q = {}
    if user['role'] != 'master':
        q['tenant_id'] = user.get('tenant_id', '')
    logs = await db.audit_logs.find(q, {"_id": 0}).sort('timestamp', -1).to_list(10000)
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    import io
    wb = Workbook()
    ws = wb.active
    ws.title = "Auditoria"
    for i, h in enumerate(['Data', 'Usuario', 'Acao', 'Entidade', 'ID', 'Detalhes'], 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
    for r, l in enumerate(logs, 2):
        ws.cell(row=r, column=1, value=l.get('timestamp', ''))
        ws.cell(row=r, column=2, value=l.get('user_email', ''))
        ws.cell(row=r, column=3, value=l.get('action', ''))
        ws.cell(row=r, column=4, value=l.get('entity_type', ''))
        ws.cell(row=r, column=5, value=l.get('entity_id', ''))
        ws.cell(row=r, column=6, value=str(l.get('changes', '') or ''))
    for col in 'ABCDEF':
        ws.column_dimensions[col].width = 22
    buf = io.BytesIO()
    wb.save(buf)
    return Response(content=buf.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=auditoria.xlsx"})

@api.get("/notifications")
async def list_notifications(user: dict = Depends(get_current_user)):
    q = {"user_id": user['sub']}
    return await db.notifications.find(q, {"_id": 0}).sort('created_at', -1).to_list(100)

@api.get("/notifications/unread-count")
async def unread_count(user: dict = Depends(get_current_user)):
    c = await db.notifications.count_documents({"user_id": user['sub'], "read": False})
    return {"count": c}

@api.patch("/notifications/{nid}/read")
async def mark_read(nid: str, user: dict = Depends(get_current_user)):
    await db.notifications.update_one({"id": nid, "user_id": user['sub']}, {"$set": {"read": True}})
    return {"message": "Lida"}

@api.post("/notifications/read-all")
async def read_all(user: dict = Depends(get_current_user)):
    await db.notifications.update_many({"user_id": user['sub'], "read": False}, {"$set": {"read": True}})
    return {"message": "Todas lidas"}

# ═══════════════════════════════════════════════
# SEED
# ═══════════════════════════════════════════════

@api.post("/seed")
async def seed():
    existing = await db.users.find_one({"email": "admin@gestaotj.com"}, {"_id": 0})
    if existing:
        return {"message": "Ja inicializado"}
    now = datetime.now(timezone.utc).isoformat()
    tenant = {"id": gen_id(), "name": "TJ Principal", "slug": "tj", "active": True, "created_at": now}
    await db.tenants.insert_one(tenant)
    users = [
        {"id": gen_id(), "email": "admin@gestaotj.com", "name": "Administrador", "role": "master", "tenant_id": "", "warehouse_id": "", "password_hash": hash_password("Admin@123456"), "active": True, "created_at": now},
        {"id": gen_id(), "email": "gerente@gestaotj.com", "name": "Gerente", "role": "admin", "tenant_id": tenant['id'], "warehouse_id": "", "password_hash": hash_password("Gerente@123"), "active": True, "created_at": now},
        {"id": gen_id(), "email": "usuario@gestaotj.com", "name": "Usuario", "role": "logistica", "tenant_id": tenant['id'], "warehouse_id": "", "password_hash": hash_password("Usuario@123"), "active": True, "created_at": now},
    ]
    await db.users.insert_many(users)
    await db.users.create_index("email", unique=True)
    return {"message": "Sistema inicializado", "tenant_id": tenant['id']}

# ═══════════════════════════════════════════════
# APP SETUP
# ═══════════════════════════════════════════════

app.include_router(api)
app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("shutdown")
async def shutdown():
    client.close()
